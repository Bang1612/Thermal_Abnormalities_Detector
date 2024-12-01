import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load the Excel file
file_path = 'Dis.xlsx'  # Replace with your file path
df = pd.read_excel(file_path, header =None)
print(df)


# ws = load_workbook('Dis.xlsx')
# # Select the active sheet (or you can specify a sheet name if needed)
# ws = ws.active

# # Define colors for the cells
# blue_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
# yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
# red_fill = PatternFill(start_color="FF4500", end_color="FF4500", fill_type="solid")

# # Loop through the DataFrame and color the cells based on values
# for row in range(1, ws.max_row + 1):
#     for col in range(1, ws.max_column + 1):
#         cell = ws.cell(row=row, column=col)
#         if cell.value == 1:  # or another range close to 35
#             cell.fill = blue_fill
#         elif cell.value == 2:  # or another range close to 40
#             cell.fill = yellow_fill
#         elif cell.value == 3:  # or another range close to 55
#             cell.fill = red_fill
# # ws.save('Dis.xlsx')

# Replace cells with specific values
df_replaced = df.copy()
# df_replaced = df_replaced.applymap(
#     lambda x: round(np.random.uniform(34.5, 35.5), 1) if x == 1
#     else round(np.random.uniform(39.5, 40.5), 1) if x == 2
#     else round(np.random.uniform(54.5, 55.5), 1) if x == 3
#     else x
# )

# # Save the updated dataframe back to an Excel file
# # df_replaced.to_csv('Updated_Dis.csv', index=False, header=False)

# # print("Replacement completed and saved to 'Updated_Dis.csv'")

# flattened_values = df_replaced.values.flatten()
# # Save the flattened values into a CSV file, with all values in one line
output_file_path = 'Updated_Dis_single_line.csv'
# # Append the flattened values as a new row in the CSV
# with open(output_file_path, mode='a', newline='') as file:
#     pd.DataFrame([flattened_values]).to_csv(file, index=False, header=False)

# print("All values appended to a new line in 'Updated_Dis_single_line.csv'")
# Run the loop 50 times
for _ in range(50):
    # Replace cells with specific values
    df_replaced = df.copy()
    df_replaced = df_replaced.applymap(
        lambda x: round(np.random.uniform(34.5, 35.5), 1) if x == 1
        else round(np.random.uniform(39.5, 40.5), 1) if x == 2
        else round(np.random.uniform(54.5, 55.5), 1) if x == 3
        else x
    )

    # Flatten the DataFrame to a single line
    flattened_values = df_replaced.values.flatten()

    # Append the flattened values as a new row in the CSV
    with open(output_file_path, mode='a', newline='') as file:
        pd.DataFrame([flattened_values]).to_csv(file, index=False, header=False)

    print(f"Iteration {_+1} completed and appended to 'Updated_Dis_single_line.csv'")

print("50 iterations completed!")
